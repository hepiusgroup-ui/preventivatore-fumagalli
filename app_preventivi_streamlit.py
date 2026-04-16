import io
import math
import re
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any

import pandas as pd
import streamlit as st
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT

# =========================
# CONFIG
# =========================
DEFAULT_CATALOGO_PDF = "FCR_Catalogo Riabilitazione (1).pdf"
st.set_page_config(
    page_title="Fumagalli Care&Reha - Preventivi",
    page_icon="📄",
    layout="wide"
)
ACCESS_PASSWORD = "fumagalli2026"

def require_login():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.title("Accesso riservato")
        password = st.text_input("Inserisci password", type="password")

        if st.button("Accedi"):
            if password == ACCESS_PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Password non corretta")

        st.stop()

require_login()
APP_TITLE = "Preventivatore Fumagalli Care&Reha"
DEFAULT_LISTINO_FILES = [
    "listino_fumagalli_riabilitazione.xlsx",
    "listino_fumagalli_contract.xlsx",
    "listino_fumagalli_child.xlsx"
]
DEFAULT_LOGO_FILES = ["logo_fumagalli.png", "Hepius.png", "logo_hepius.png"]

DEFAULT_CATALOGO_PDF = "FCR_Catalogo Riabilitazione (1).pdf"
DEFAULT_MAPPA_CATALOGO = "mappa_catalogo.xlsx"

BRAND_BLUE = "1F4E78"
BRAND_LIGHT = "DCE6F1"
BRAND_GREY = "EDEDED"
BRAND_GREEN = "E2F0D9"
WHITE = "FFFFFF"

REQUIRED_LOGICAL_FIELDS = {
    "codice": ["codice", "code", "cod.articolo", "codice articolo", "article code", "sku", "part number"],
    "descrizione": ["descrizione", "description", "descrizione articolo", "articolo", "prodotto", "item", "nome"],
    "prezzo": ["prezzo", "price", "prezzo pubblico", "listino", "public price", "prezzo ex iva", "prezzo iva esclusa"],
    "foto": ["foto", "immagine", "image", "img", "image path", "percorso immagine", "picture"],
}


# =========================
# UTILS
# =========================

def clean_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def find_column(columns: List[str], aliases: List[str]) -> Optional[str]:
    normalized = {clean_name(c): c for c in columns}
    for alias in aliases:
        key = clean_name(alias)
        if key in normalized:
            return normalized[key]
    return None


def normalize_dataframe(df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    codice_col = find_column(list(df.columns), REQUIRED_LOGICAL_FIELDS["codice"])
    descr_col = find_column(list(df.columns), REQUIRED_LOGICAL_FIELDS["descrizione"])
    prezzo_col = find_column(list(df.columns), REQUIRED_LOGICAL_FIELDS["prezzo"])
    foto_col = find_column(list(df.columns), REQUIRED_LOGICAL_FIELDS["foto"])

    if not codice_col or not descr_col:
        raise ValueError(
            f"Nel file {source_name} non trovo colonne compatibili con codice e descrizione."
        )

    out = pd.DataFrame()
    out["codice"] = df[codice_col].astype(str).str.strip()
    out["descrizione"] = df[descr_col].astype(str).str.strip()

    if prezzo_col:
        out["prezzo"] = pd.to_numeric(df[prezzo_col], errors="coerce")
    else:
        out["prezzo"] = pd.NA

    if foto_col:
        out["foto"] = df[foto_col].astype(str).replace("nan", "").fillna("")
    else:
        out["foto"] = ""

    out["sorgente"] = source_name

    # 🔥 NOMI LISTINI USER-FRIENDLY
    out["sorgente"] = out["sorgente"].replace({
        "listino_fumagalli_riabilitazione.xlsx": "Riabilitazione",
        "listino_fumagalli_contract.xlsx": "Contract",
        "listino_fumagalli_child.xlsx": "Child"
    })

    out = out[(out["codice"] != "") & (out["descrizione"] != "")]
    out = out.drop_duplicates(subset=["codice", "descrizione"], keep="first").reset_index(drop=True)

    return out


def load_price_lists(uploaded_files: List[Any]) -> pd.DataFrame:
    frames = []

    if uploaded_files:
        for file in uploaded_files:
            df = pd.read_excel(file)
            frames.append(normalize_dataframe(df, getattr(file, "name", "listino_caricato")))
    else:
        found = []
        for file_name in DEFAULT_LISTINO_FILES:
            p = Path(file_name)
            if p.exists():
                found.append(p)
        if not found:
            raise FileNotFoundError(
                "Non trovo listino_fumagalli.xlsx e/o listino_hepius.xlsx nella stessa cartella dell'app. "
                "Caricali dalla sidebar."
            )
        for p in found:
            df = pd.read_excel(p)
            frames.append(normalize_dataframe(df, p.name))

    if not frames:
        raise ValueError("Nessun listino disponibile.")
    all_df = pd.concat(frames, ignore_index=True)
    all_df["search_text"] = (
        all_df["codice"].fillna("").astype(str).str.lower() + " " +
        all_df["descrizione"].fillna("").astype(str).str.lower()
    )
    return all_df


def smart_search(query: str, df: pd.DataFrame, limit: int = 20) -> pd.DataFrame:
    query = query.strip().lower()
    if not query:
        return df.head(limit).copy()

    tokens = [t for t in re.split(r"\s+", query) if t]
    scores = []

    for idx, row in df.iterrows():
        text = row["search_text"]
        score_1 = fuzz.token_set_ratio(query, text)
        score_2 = fuzz.partial_ratio(query, text)
        token_bonus = sum(7 for t in tokens if t in text)
        score = (score_1 * 0.65) + (score_2 * 0.35) + token_bonus
        scores.append((idx, score))

    scored_idx = sorted(scores, key=lambda x: x[1], reverse=True)[:limit]
    result = df.loc[[i for i, _ in scored_idx]].copy()
    result["match_score"] = [round(s, 1) for _, s in scored_idx]
    return result.reset_index(drop=True)


def money(value: float) -> str:
    try:
        return f"EUR {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "EUR 0,00"


def parse_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("CHF", "").replace("€", "").replace("%", "")
    text = text.replace(".", "").replace(",", ".") if text.count(",") == 1 and text.count(".") >= 1 else text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        return default


def find_existing_logo_candidates() -> Dict[str, Optional[Path]]:
    results = {"fumagalli": None, "hepius": None}
    for p in Path(".").iterdir():
        if p.is_file():
            name = p.name.lower()
            if ("fumagalli" in name or "care&reha" in name or "care" in name) and p.suffix.lower() in [".png", ".jpg", ".jpeg"]:
                results["fumagalli"] = p
            if "hepius" in name and p.suffix.lower() in [".png", ".jpg", ".jpeg"]:
                results["hepius"] = p
    return results


def safe_add_image(ws, img_path: Optional[Path], cell: str, width_px: int = 180):
    if not img_path or not Path(img_path).exists():
        return
    try:
        img = XLImage(str(img_path))
        original_width = img.width or width_px
        ratio = width_px / original_width
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
        ws.add_image(img, cell)
    except Exception:
        return


def resolve_logo_file(uploaded_logo, fallback_path: Optional[Path], tmp_name: str) -> Optional[Path]:
    if uploaded_logo is not None:
        out = Path(tmp_name)
        out.write_bytes(uploaded_logo.getbuffer())
        return out
    if fallback_path and fallback_path.exists():
        return fallback_path
    return None


def try_get_photo_path(photo_value: Any) -> Optional[Path]:
    if photo_value is None:
        return None
    s = str(photo_value).strip()
    if not s or s.lower() == "nan":
        return None
    p = Path(s)
    return p if p.exists() else None


def build_quote_excel(customer_data: Dict[str, str], items_df: pd.DataFrame, extra_discount_pct: float,
                      fumagalli_logo: Optional[Path], hepius_logo: Optional[Path]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Preventivo"
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = {
        "A": 12, "B": 16, "C": 48, "D": 14, "E": 14, "F": 14, "G": 14, "H": 18
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header area
    ws.merge_cells("C1:H1")
    ws["C1"] = "PREVENTIVO"
    ws["C1"].font = Font(size=18, bold=True, color=BRAND_BLUE)
    ws["C1"].alignment = Alignment(horizontal="right", vertical="center")

    ws["G2"] = "Data"
    ws["H2"] = datetime.now().strftime("%d/%m/%Y")
    ws["G3"] = "Numero"
    ws["H3"] = datetime.now().strftime("OF-%Y%m%d-%H%M")

    for cell in ["G2", "G3"]:
        ws[cell].font = Font(bold=True, color=BRAND_BLUE)

    safe_add_image(ws, fumagalli_logo, "A1", width_px=170)
    safe_add_image(ws, hepius_logo, "A4", width_px=150)

    # Customer block
    top = 8
    ws.merge_cells(start_row=top, start_column=1, end_row=top, end_column=8)
    ws.cell(top, 1).value = "DATI CLIENTE"
    ws.cell(top, 1).fill = PatternFill("solid", fgColor=BRAND_BLUE)
    ws.cell(top, 1).font = Font(color=WHITE, bold=True)
    ws.cell(top, 1).alignment = Alignment(horizontal="left")

    fields = [
        ("Cliente", customer_data.get("cliente", "")),
        ("Attenzione", customer_data.get("contatto", "")),
        ("Indirizzo", customer_data.get("indirizzo", "")),
        ("CAP / Città", customer_data.get("cap_citta", "")),
        ("Email", customer_data.get("email", "")),
        ("Telefono", customer_data.get("telefono", "")),
        ("Oggetto", customer_data.get("oggetto", "")),
    ]
    r = top + 1
    for label, value in fields:
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = Font(bold=True, color=BRAND_BLUE)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(r, 2).value = value
        r += 1

    table_row = r + 2
    headers = ["Foto", "Codice", "Descrizione", "Q.tà", "Prezzo unit.", "Sconto riga", "Netto riga", "Origine"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(table_row, col_idx)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color=WHITE)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    currency_fmt = '€ #,##0.00'
    pct_fmt = '0.0%'

    data_start = table_row + 1
    current_row = data_start

    for _, item in items_df.iterrows():
        ws.row_dimensions[current_row].height = 58

        ws.cell(current_row, 2).value = item["codice"]
        ws.cell(current_row, 3).value = item["descrizione"]
        ws.cell(current_row, 4).value = float(item["quantita"])
        ws.cell(current_row, 5).value = float(item["prezzo_unitario"])
        ws.cell(current_row, 6).value = float(item["sconto_riga_pct"]) / 100.0
        ws.cell(current_row, 7).value = f"=D{current_row}*E{current_row}*(1-F{current_row})"
        ws.cell(current_row, 8).value = item["sorgente"]

        ws.cell(current_row, 4).alignment = Alignment(horizontal="center")
        ws.cell(current_row, 5).number_format = currency_fmt
        ws.cell(current_row, 6).number_format = pct_fmt
        ws.cell(current_row, 7).number_format = currency_fmt

        for c in range(1, 9):
            cell = ws.cell(current_row, c)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            border = Side(style="thin", color="D9D9D9")
            cell.border = Border(left=border, right=border, top=border, bottom=border)

        photo_path = try_get_photo_path(item.get("foto", ""))
        if photo_path:
            safe_add_image(ws, photo_path, f"A{current_row}", width_px=62)

        current_row += 1

    subtotal_row = current_row + 1
    extra_row = subtotal_row + 1
    total_row = subtotal_row + 2

    ws.cell(subtotal_row, 6).value = "Subtotale"
    ws.cell(subtotal_row, 6).font = Font(bold=True, color=BRAND_BLUE)
    ws.cell(subtotal_row, 7).value = f"=SUM(G{data_start}:G{current_row-1})"
    ws.cell(subtotal_row, 7).number_format = currency_fmt

    ws.cell(extra_row, 6).value = "Sconto extra offerta"
    ws.cell(extra_row, 6).font = Font(bold=True, color=BRAND_BLUE)
    ws.cell(extra_row, 7).value = extra_discount_pct / 100.0
    ws.cell(extra_row, 7).number_format = pct_fmt

    ws.cell(total_row, 6).value = "Totale offerta"
    ws.cell(total_row, 6).fill = PatternFill("solid", fgColor=BRAND_GREEN)
    ws.cell(total_row, 6).font = Font(bold=True, color=BRAND_BLUE)
    ws.cell(total_row, 7).value = f"=G{subtotal_row}*(1-G{extra_row})"
    ws.cell(total_row, 7).fill = PatternFill("solid", fgColor=BRAND_GREEN)
    ws.cell(total_row, 7).font = Font(bold=True, color=BRAND_BLUE)
    ws.cell(total_row, 7).number_format = currency_fmt

    note_row = total_row + 3
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    ws.cell(note_row, 1).value = (
        "Prezzi espressi in EURO. Trasporto, montaggio ed IVA esclusa salvo diversa indicazione. "
        "Sconti di riga ed eventuale sconto finale sono cumulativi."
    )
    ws.cell(note_row, 1).font = Font(italic=True, color="666666")

    footer_row = note_row + 3
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    ws.cell(footer_row, 1).value = "Fumagalli Care&Reha Srl"
    ws.cell(footer_row, 1).font = Font(bold=True, color=BRAND_BLUE)
    ws.cell(footer_row, 1).alignment = Alignment(horizontal="center")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
def build_quote_pdf(
    customer_data: Dict[str, str],
    items_df: pd.DataFrame,
    extra_discount_pct: float,
    fumagalli_logo: Optional[Path],
    hepius_logo: Optional[Path]
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleCustom",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        textColor=colors.HexColor("#1F4E78"),
        alignment=TA_RIGHT,
        spaceAfter=6
    )

    label_style = ParagraphStyle(
        "Label",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=2
    )

    value_style = ParagraphStyle(
        "Value",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.black,
        leading=11,
        spaceAfter=3
    )

    note_style = ParagraphStyle(
        "Note",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=8,
        textColor=colors.grey,
        leading=10
    )

    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10
    )

    cell_bold_style = ParagraphStyle(
        "CellBoldStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white
    )

    story = []

    # =========================
    # HEADER
    # =========================
    left_parts = []

    if fumagalli_logo and Path(fumagalli_logo).exists():
        try:
            left_parts.append(RLImage(str(fumagalli_logo), width=42 * mm, height=16 * mm))
            left_parts.append(Spacer(1, 2 * mm))
        except Exception:
            pass

    if hepius_logo and Path(hepius_logo).exists():
        try:
            left_parts.append(RLImage(str(hepius_logo), width=35 * mm, height=14 * mm))
        except Exception:
            pass

    if not left_parts:
        left_parts = [Paragraph("Fumagalli Care&Reha Srl", value_style)]

    right_cell = [
        Paragraph("PREVENTIVO", title_style),
        Paragraph(f"<b>Data:</b> {datetime.now().strftime('%d/%m/%Y')}", value_style),
        Paragraph(f"<b>Numero:</b> OF-{datetime.now().strftime('%Y%m%d-%H%M')}", value_style),
    ]

    header_table = Table(
        [[left_parts, right_cell]],
        colWidths=[78 * mm, 92 * mm]
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6 * mm))

    # =========================
    # DATI CLIENTE
    # =========================
    story.append(Paragraph("DATI CLIENTE", label_style))

    cliente_rows = [
        [Paragraph("<b>Cliente</b>", value_style), Paragraph(str(customer_data.get("cliente", "")), value_style)],
        [Paragraph("<b>Attenzione</b>", value_style), Paragraph(str(customer_data.get("contatto", "")), value_style)],
        [Paragraph("<b>Indirizzo</b>", value_style), Paragraph(str(customer_data.get("indirizzo", "")), value_style)],
        [Paragraph("<b>CAP / Città</b>", value_style), Paragraph(str(customer_data.get("cap_citta", "")), value_style)],
        [Paragraph("<b>Email</b>", value_style), Paragraph(str(customer_data.get("email", "")), value_style)],
        [Paragraph("<b>Telefono</b>", value_style), Paragraph(str(customer_data.get("telefono", "")), value_style)],
        [Paragraph("<b>Oggetto</b>", value_style), Paragraph(str(customer_data.get("oggetto", "")), value_style)],
    ]

    cliente_table = Table(cliente_rows, colWidths=[38 * mm, 132 * mm])
    cliente_table.setStyle(TableStyle([
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1F4E78")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9D9D9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(cliente_table)
    story.append(Spacer(1, 6 * mm))

    # =========================
    # TABELLA ARTICOLI
    # =========================
    table_data = [[
        Paragraph("<b>Codice</b>", cell_bold_style),
        Paragraph("<b>Descrizione</b>", cell_bold_style),
        Paragraph("<b>Q.tà</b>", cell_bold_style),
        Paragraph("<b>Prezzo unit.</b>", cell_bold_style),
        Paragraph("<b>Sconto %</b>", cell_bold_style),
        Paragraph("<b>Netto riga</b>", cell_bold_style),
    ]]

    subtotal = 0.0
    gross = 0.0

    for _, item in items_df.iterrows():
        qty = float(item["quantita"])
        price = float(item["prezzo_unitario"])
        discount = float(item["sconto_riga_pct"])
        row_total = qty * price * (1 - discount / 100.0)

        gross += qty * price
        subtotal += row_total

        codice_p = Paragraph(str(item["codice"]), cell_style)
        descr_p = Paragraph(str(item["descrizione"]), cell_style)
        qty_p = Paragraph(f"{qty:,.0f}".replace(",", "."), cell_style)
        price_p = Paragraph(money(price), cell_style)
        discount_p = Paragraph(f"{discount:.1f}%", cell_style)
        row_total_p = Paragraph(money(row_total), cell_style)

        table_data.append([
            codice_p,
            descr_p,
            qty_p,
            price_p,
            discount_p,
            row_total_p
        ])

    total_final = subtotal * (1 - extra_discount_pct / 100.0)

    offer_table = Table(
        table_data,
        colWidths=[26 * mm, 80 * mm, 14 * mm, 23 * mm, 18 * mm, 24 * mm],
        repeatRows=1
    )
    offer_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9D9D9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(offer_table)
    story.append(Spacer(1, 6 * mm))

    # =========================
    # TOTALI
    # =========================
    totals_data = [
        [Paragraph("<b>Totale lordo</b>", value_style), Paragraph(money(gross), value_style)],
        [Paragraph("<b>Subtotale dopo sconti riga</b>", value_style), Paragraph(money(subtotal), value_style)],
        [Paragraph("<b>Sconto extra finale</b>", value_style), Paragraph(f"{extra_discount_pct:.1f}%", value_style)],
        [Paragraph("<b>Totale finale offerta escluso IVA</b>", value_style), Paragraph(money(total_final), value_style)],
    ]

    totals_table = Table(totals_data, colWidths=[58 * mm, 34 * mm], hAlign="RIGHT")
    totals_table.setStyle(TableStyle([
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1F4E78")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E2F0D9")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9D9D9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 6 * mm))

    # =========================
    # NOTE FINALI
    # =========================
    story.append(Paragraph(
        "Prezzi espressi in EURO, IVA e costi di spedizione esclusi salvo diversa indicazione. "
        "Sconti di riga ed eventuale sconto finale sono cumulativi.",
        note_style
    ))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph("Fumagalli Care&Reha Srl", label_style))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

    # Header con loghi + titolo
    header_data = []

    left_parts = []
    if fumagalli_logo and Path(fumagalli_logo).exists():
        left_parts.append(RLImage(str(fumagalli_logo), width=42 * mm, height=16 * mm))
    if hepius_logo and Path(hepius_logo).exists():
        left_parts.append(RLImage(str(hepius_logo), width=35 * mm, height=14 * mm))

    left_cell = left_parts if left_parts else [Paragraph("Fumagalli Care&Reha", value_style)]
    right_cell = [
        Paragraph("PREVENTIVO", title_style),
        Paragraph(f"<b>Data:</b> {datetime.now().strftime('%d/%m/%Y')}", value_style),
        Paragraph(f"<b>Numero:</b> OF-{datetime.now().strftime('%Y%m%d-%H%M')}", value_style),
    ]

    header_table = Table([[left_cell, right_cell]], colWidths=[80 * mm, 90 * mm])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 8 * mm))

    # Dati cliente
    story.append(Paragraph("DATI CLIENTE", label_style))

    cliente_rows = [
        ["Cliente", customer_data.get("cliente", "")],
        ["Attenzione", customer_data.get("contatto", "")],
        ["Indirizzo", customer_data.get("indirizzo", "")],
        ["CAP / Città", customer_data.get("cap_citta", "")],
        ["Email", customer_data.get("email", "")],
        ["Telefono", customer_data.get("telefono", "")],
        ["Oggetto", customer_data.get("oggetto", "")],
    ]

    cliente_table = Table(cliente_rows, colWidths=[35 * mm, 135 * mm])
    cliente_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.white),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1F4E78")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9D9D9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(cliente_table)
    story.append(Spacer(1, 8 * mm))

    # Tabella articoli
    table_data = [[
        "Codice", "Descrizione", "Q.tà", "Prezzo unit.", "Sconto %", "Netto riga"
    ]]

    subtotal = 0.0
    gross = 0.0

    for _, item in items_df.iterrows():
        qty = float(item["quantita"])
        price = float(item["prezzo_unitario"])
        discount = float(item["sconto_riga_pct"])
        row_total = qty * price * (1 - discount / 100.0)

        gross += qty * price
        subtotal += row_total

        table_data.append([
            str(item["codice"]),
            str(item["descrizione"]),
            f"{qty:,.0f}".replace(",", "."),
            money(price),
            f"{discount:.1f}%",
            money(row_total)
        ])

    total_final = subtotal * (1 - extra_discount_pct / 100.0)

    offer_table = Table(
        table_data,
        colWidths=[25 * mm, 78 * mm, 14 * mm, 24 * mm, 18 * mm, 24 * mm],
        repeatRows=1
    )
    offer_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9D9D9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(offer_table)
    story.append(Spacer(1, 6 * mm))

    # Totali
    totals_data = [
        ["Totale lordo", money(gross)],
        ["Subtotale dopo sconti riga", money(subtotal)],
        ["Sconto extra finale", f"{extra_discount_pct:.1f}%"],
        ["Totale finale offerta", money(total_final)],
    ]
    totals_table = Table(totals_data, colWidths=[55 * mm, 35 * mm], hAlign="RIGHT")
    totals_table.setStyle(TableStyle([
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1F4E78")),
        ("FONTNAME", (0, 0), (-1, -2), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E2F0D9")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9D9D9")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 6 * mm))

    # Note finali
    story.append(Paragraph(
        "Prezzi espressi in EURO, IVA e costi di spedizione esclusi salvo diversa indicazione. "
        "Sconti di riga ed eventuale sconto finale sono cumulativi.",
        note_style
    ))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("Fumagalli Care&Reha Srl", label_style))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
def merge_main_pdf_with_catalogs(main_pdf_bytes: bytes, catalog_pdf_list: List[bytes]) -> bytes:
    writer = PdfWriter()

    if main_pdf_bytes:
        main_reader = PdfReader(io.BytesIO(main_pdf_bytes))
        for page in main_reader.pages:
            writer.add_page(page)

    for pdf_bytes in catalog_pdf_list:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        for page in reader.pages:
            writer.add_page(page)

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output.getvalue()
def compute_totals(items: List[Dict[str, Any]], extra_discount_pct: float) -> Dict[str, float]:
    subtotal = 0.0
    gross = 0.0
    for item in items:
        qty = parse_float(item["quantita"], 1.0)
        price = parse_float(item["prezzo_unitario"], 0.0)
        row_disc = parse_float(item["sconto_riga_pct"], 0.0)
        gross += qty * price
        subtotal += qty * price * (1 - row_disc / 100.0)
    final_total = subtotal * (1 - extra_discount_pct / 100.0)
    return {
        "lordo": gross,
        "subtotale": subtotal,
        "totale_finale": final_total,
        "sconto_totale": gross - final_total,
    }
def load_catalog_mapping(mapping_file: str = DEFAULT_MAPPA_CATALOGO) -> pd.DataFrame:
    p = Path(mapping_file)
    if not p.exists():
        raise FileNotFoundError(f"Non trovo il file di mapping catalogo: {mapping_file}")

    df = pd.read_excel(p)

    required_cols = [
        "codice",
        "chiave_articolo",
        "sezione",
        "pagina_sezione",
        "pagina_inizio",
        "pagina_fine",
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Nel mapping catalogo mancano le colonne: {missing}")

    df["codice"] = df["codice"].astype(str).str.strip().str.upper()
    df["chiave_articolo"] = df["chiave_articolo"].astype(str).str.strip().str.upper()
    df["sezione"] = df["sezione"].astype(str).str.strip()

    df["pagina_sezione"] = pd.to_numeric(df["pagina_sezione"], errors="coerce")
    df["pagina_inizio"] = pd.to_numeric(df["pagina_inizio"], errors="coerce")
    df["pagina_fine"] = pd.to_numeric(df["pagina_fine"], errors="coerce")

    df = df.dropna(subset=["pagina_sezione", "pagina_inizio", "pagina_fine"]).copy()

    df["pagina_sezione"] = df["pagina_sezione"].astype(int)
    df["pagina_inizio"] = df["pagina_inizio"].astype(int)
    df["pagina_fine"] = df["pagina_fine"].astype(int)

    return df


def find_catalog_rows_for_items(items_df: pd.DataFrame, mapping_df: pd.DataFrame) -> pd.DataFrame:
    found_rows = []

    for _, item in items_df.iterrows():
        descr = str(item.get("descrizione", "")).strip().upper()
        codice = str(item.get("codice", "")).strip().upper()

        matches = pd.DataFrame()

        # 1. match per codice
        if "codice" in mapping_df.columns:
            matches = mapping_df[
                mapping_df["codice"].astype(str).str.strip().str.upper() == codice
            ].copy()

        # 2. fallback: chiave_articolo uguale alla descrizione
        if matches.empty:
            matches = mapping_df[
                mapping_df["chiave_articolo"].astype(str).str.strip().str.upper() == descr
            ].copy()

        # 3. fallback: chiave_articolo contenuta nella descrizione
        if matches.empty:
            matches = mapping_df[
                mapping_df["chiave_articolo"].astype(str).str.strip().str.upper().apply(
                    lambda x: x in descr if isinstance(x, str) and x != "" else False
                )
            ].copy()

        if not matches.empty:
            row = matches.iloc[0].copy()
            row["codice_offerta"] = codice
            row["descrizione_offerta"] = descr
            found_rows.append(row)

    if not found_rows:
        return pd.DataFrame(columns=mapping_df.columns)

    return pd.DataFrame(found_rows).drop_duplicates().reset_index(drop=True)
def build_catalog_pdf_for_single_item(
    item_row: pd.Series,
    catalog_pdf_path: str = DEFAULT_CATALOGO_PDF,
    mapping_file: str = DEFAULT_MAPPA_CATALOGO
) -> bytes:
    catalog_path = Path(catalog_pdf_path)
    if not catalog_path.exists():
        raise FileNotFoundError(f"Non trovo il catalogo PDF: {catalog_pdf_path}")

    mapping_df = load_catalog_mapping(mapping_file)
    item_df = pd.DataFrame([item_row])
    selected_rows = find_catalog_rows_for_items(item_df, mapping_df)

    if selected_rows.empty:
        raise ValueError(f"Nessuna corrispondenza trovata per articolo {item_row.get('codice', '')}")

    row = selected_rows.iloc[0]

    reader = PdfReader(str(catalog_path))
    writer = PdfWriter()
    added_pages = set()

    pagina_sezione = int(row["pagina_sezione"])
    pagina_inizio = int(row["pagina_inizio"])
    pagina_fine = int(row["pagina_fine"])

    # pagina introduttiva sezione
    if 1 <= pagina_sezione <= len(reader.pages) and pagina_sezione not in added_pages:
        writer.add_page(reader.pages[pagina_sezione - 1])
        added_pages.add(pagina_sezione)

    # pagine prodotto
    for page_num in range(pagina_inizio, pagina_fine + 1):
        if 1 <= page_num <= len(reader.pages) and page_num not in added_pages:
            writer.add_page(reader.pages[page_num - 1])
            added_pages.add(page_num)

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output.getvalue()

def build_catalog_attachment_pdf(
    items_df: pd.DataFrame,
    catalog_pdf_path: str = DEFAULT_CATALOGO_PDF,
    mapping_file: str = DEFAULT_MAPPA_CATALOGO
) -> bytes:
    catalog_path = Path(catalog_pdf_path)
    if not catalog_path.exists():
        raise FileNotFoundError(f"Non trovo il catalogo PDF: {catalog_pdf_path}")

    mapping_df = load_catalog_mapping(mapping_file)
    selected_rows = find_catalog_rows_for_items(items_df, mapping_df)

    if selected_rows.empty:
        raise ValueError("Nessuna corrispondenza trovata tra articoli offerta e mappa catalogo.")

    reader = PdfReader(str(catalog_path))
    writer = PdfWriter()

    added_pages = set()

    # 1. aggiungi prima le pagine sezione uniche
    section_pages = sorted(selected_rows["pagina_sezione"].drop_duplicates().tolist())
    for page_num in section_pages:
        page_index = page_num - 1
        if 0 <= page_index < len(reader.pages) and page_num not in added_pages:
            writer.add_page(reader.pages[page_index])
            added_pages.add(page_num)

    # 2. aggiungi poi le pagine prodotto
    for _, row in selected_rows.iterrows():
        start_page = int(row["pagina_inizio"])
        end_page = int(row["pagina_fine"])

        for page_num in range(start_page, end_page + 1):
            page_index = page_num - 1
            if 0 <= page_index < len(reader.pages) and page_num not in added_pages:
                writer.add_page(reader.pages[page_index])
                added_pages.add(page_num)

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output.getvalue()

# =========================
# SESSION STATE
# =========================

if "cart_items" not in st.session_state:
    st.session_state.cart_items = []

if "search_results_cache" not in st.session_state:
    st.session_state.search_results_cache = pd.DataFrame()

if "last_selected_code" not in st.session_state:
    st.session_state.last_selected_code = None

if "qty_add" not in st.session_state:
    st.session_state.qty_add = 1.0

if "disc_add" not in st.session_state:
    st.session_state.disc_add = 0.0

if "price_add" not in st.session_state:
    st.session_state.price_add = 0.0


# =========================
# UI
# =========================

st.title(APP_TITLE)
st.caption("Ricerca intelligente da 2 listini, prezzi mancanti gestibili manualmente, sconto per riga e sconto extra finale.")

with st.sidebar:
    st.header("Dati sorgente")
    uploaded_lists = st.file_uploader(
        "Carica 1 o 2 listini Excel",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="Se non carichi nulla, l'app prova a leggere listino_fumagalli.xlsx e listino_hepius.xlsx dalla stessa cartella."
    )
    st.markdown("---")
    st.subheader("Loghi")
    uploaded_logo_fum = st.file_uploader("Logo Fumagalli", type=["png", "jpg", "jpeg"], key="fum_logo")
    uploaded_logo_hep = st.file_uploader("Logo Hepius", type=["png", "jpg", "jpeg"], key="hep_logo")

logo_fallbacks = find_existing_logo_candidates()
fum_logo_path = resolve_logo_file(uploaded_logo_fum, logo_fallbacks.get("fumagalli"), "_logo_fumagalli_tmp.png")
hep_logo_path = resolve_logo_file(uploaded_logo_hep, logo_fallbacks.get("hepius"), "_logo_hepius_tmp.png")

try:
    master_df = load_price_lists(uploaded_lists)
    st.success(f"Listino caricato: {len(master_df)} articoli disponibili.")
except Exception as e:
    st.error(str(e))
    st.stop()

# Customer form
with st.expander("Dati cliente e testata offerta", expanded=True):
    col1, col2, col3 = st.columns(3)
    with col1:
        cliente = st.text_input("Cliente / Ragione sociale")
        contatto = st.text_input("Attenzione")
        telefono = st.text_input("Telefono")
    with col2:
        indirizzo = st.text_input("Indirizzo")
        cap_citta = st.text_input("CAP / Città")
        email = st.text_input("Email")
    with col3:
        oggetto = st.text_input("Oggetto offerta", value="Fornitura ausili / dispositivi")
        note = st.text_area("Note interne", value="", height=90)

st.markdown("## Ricerca articoli")

available_sources = sorted(master_df["sorgente"].dropna().unique().tolist())

f_col1, f_col2 = st.columns([3, 2])
with f_col1:
    query = st.text_input("Parole chiave", placeholder="es. carrozzina pieghevole leggera")
with f_col2:
    selected_sources = st.multiselect(
        "Filtra per listino",
        options=available_sources,
        default=available_sources
    )

s_col1, s_col2 = st.columns([4, 1])
with s_col1:
    st.write("")
with s_col2:
    max_results = st.number_input("Max risultati", min_value=5, max_value=50, value=10, step=5)

if st.button("Cerca articoli", type="primary"):
    filtered_df = master_df.copy()

    if selected_sources:
        filtered_df = filtered_df[filtered_df["sorgente"].isin(selected_sources)].copy()
    else:
        filtered_df = pd.DataFrame(columns=master_df.columns)

    if not filtered_df.empty:
        st.session_state.search_results_cache = smart_search(query, filtered_df, limit=int(max_results))
    else:
        st.session_state.search_results_cache = pd.DataFrame()

    st.session_state.last_selected_code = None

    if "selected_code" in st.session_state:
        del st.session_state["selected_code"]

results_df = st.session_state.search_results_cache

if not selected_sources:
    st.warning("Seleziona almeno un listino per effettuare la ricerca.")
    results_df = pd.DataFrame()

if not results_df.empty:
    display_df = results_df[["codice", "descrizione", "prezzo", "sorgente", "match_score"]].copy()
    display_df["prezzo"] = display_df["prezzo"].apply(lambda x: "" if pd.isna(x) else float(x))
    st.dataframe(display_df, use_container_width=True, hide_index=True)

    code_to_desc = (
    results_df[["codice", "descrizione"]]
    .drop_duplicates(subset=["codice"])
    .set_index("codice")["descrizione"]
    .to_dict()
)

selected_code = st.selectbox(
    "Seleziona articolo da aggiungere",
    options=results_df["codice"].tolist(),
    format_func=lambda code: f"{code} - {code_to_desc.get(code, code)}",
    key="selected_code"
)

    selected_row = results_df.loc[results_df["codice"] == selected_code].iloc[0]
    current_price = 0.0 if pd.isna(selected_row["prezzo"]) else float(selected_row["prezzo"])

    if st.session_state.last_selected_code != selected_code:
        st.session_state.qty_add = 1.0
        st.session_state.disc_add = 0.0
        st.session_state.price_add = current_price
        st.session_state.last_selected_code = selected_code

    add1, add2, add3, add4 = st.columns([1, 1, 1, 1.4])
    with add1:
        qty_input = st.number_input(
            "Quantità",
            min_value=1.0,
            step=1.0,
            key="qty_add"
        )
    with add2:
        row_discount_input = st.number_input(
            "Sconto riga %",
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            key="disc_add"
        )
    with add3:
        manual_price_input = st.number_input(
            "Prezzo unit. ex IVA",
            min_value=0.0,
            step=1.0,
            key="price_add",
            help="Se il prezzo manca nel listino, inseriscilo manualmente qui."
        )
    with add4:
        st.write("")
        st.write("")
        if st.button("Aggiungi al preventivo"):
            st.session_state.cart_items.append({
                "codice": selected_row["codice"],
                "descrizione": selected_row["descrizione"],
                "quantita": float(qty_input),
                "prezzo_unitario": float(manual_price_input),
                "sconto_riga_pct": float(row_discount_input),
                "foto": selected_row.get("foto", ""),
                "sorgente": selected_row.get("sorgente", ""),
            })
            st.success(f"Aggiunto: {selected_row['codice']}")

st.markdown("## Preventivo in costruzione")

if st.session_state.cart_items:
    editable_df = pd.DataFrame(st.session_state.cart_items)
    editable_df = st.data_editor(
        editable_df,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        column_config={
            "codice": st.column_config.TextColumn("Codice", disabled=True),
            "descrizione": st.column_config.TextColumn("Descrizione", width="large"),
            "quantita": st.column_config.NumberColumn("Q.tà", min_value=0.0, step=1.0),
            "prezzo_unitario": st.column_config.NumberColumn("Prezzo unit. ex IVA", min_value=0.0, step=1.0),
            "sconto_riga_pct": st.column_config.NumberColumn("Sconto riga %", min_value=0.0, max_value=100.0, step=0.5),
            "foto": st.column_config.TextColumn("Foto", help="Percorso file immagine, se disponibile"),
            "sorgente": st.column_config.TextColumn("Sorgente", disabled=True),
        },
        key="cart_editor"
    )
    st.session_state.cart_items = editable_df.to_dict(orient="records")

    c1, c2, c3 = st.columns([1, 1, 1])
    with c1:
        extra_discount_pct = st.number_input(
            "Sconto extra finale %",
            min_value=0.0,
            max_value=100.0,
            value=0.0,
            step=0.5
        )
    with c2:
        if st.button("Svuota preventivo"):
            st.session_state.cart_items = []
            st.rerun()
    with c3:
        if st.button("Rimuovi ultima riga"):
            st.session_state.cart_items = st.session_state.cart_items[:-1]
            st.rerun()

    totals = compute_totals(st.session_state.cart_items, extra_discount_pct)
    k1, k2, k3 = st.columns(3)
    k1.metric("Totale lordo", money(totals["lordo"]))
    k2.metric("Subtotale dopo sconti riga", money(totals["subtotale"]))
    k3.metric("Totale finale offerta", money(totals["totale_finale"]))

    export_df = pd.DataFrame(st.session_state.cart_items)
    customer_data = {
        "cliente": cliente,
        "contatto": contatto,
        "indirizzo": indirizzo,
        "cap_citta": cap_citta,
        "email": email,
        "telefono": telefono,
        "oggetto": oggetto,
        "note": note,
    }

    excel_bytes = build_quote_excel(
        customer_data=customer_data,
        items_df=export_df,
        extra_discount_pct=float(extra_discount_pct),
        fumagalli_logo=fum_logo_path,
        hepius_logo=hep_logo_path
    )

    main_pdf_bytes = build_quote_pdf(
        customer_data=customer_data,
        items_df=export_df,
        extra_discount_pct=float(extra_discount_pct),
        fumagalli_logo=fum_logo_path,
        hepius_logo=hep_logo_path
    )

    file_name = f"preventivo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    pdf_file_name = f"preventivo_cliente_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    dl1, dl2, dl3 = st.columns(3)

    with dl1:
        st.download_button(
            label="Scarica preventivo Excel",
            data=excel_bytes,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with dl2:
        st.download_button(
            label="Scarica PDF preventivo",
            data=main_pdf_bytes,
            file_name=pdf_file_name,
            mime="application/pdf"
        )

    with dl3:
        genera_pdf_cliente_completo = st.checkbox(
            "PDF cliente completo con allegati catalogo",
            value=True,
            help="Unisce preventivo PDF e schede catalogo prodotti in un unico file."
        )

    if genera_pdf_cliente_completo:
        catalog_pdfs = []
        unmatched_codes = []

        for _, row in export_df.iterrows():
            codice = str(row.get("codice", "")).strip()
            try:
                single_pdf = build_catalog_pdf_for_single_item(
                    item_row=row,
                    catalog_pdf_path=DEFAULT_CATALOGO_PDF,
                    mapping_file=DEFAULT_MAPPA_CATALOGO
                )
                catalog_pdfs.append(single_pdf)
            except Exception:
                unmatched_codes.append(codice)

        if catalog_pdfs:
            schede_tecniche_pdf = merge_main_pdf_with_catalogs(b"", catalog_pdfs)

            st.download_button(
                label="Scarica PDF schede tecniche",
                data=schede_tecniche_pdf,
                file_name=f"schede_tecniche_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf"
            )

        if unmatched_codes:
            st.warning(
                "Attenzione: nessun allegato catalogo trovato per i seguenti articoli: "
                + ", ".join(unmatched_codes)
            )
